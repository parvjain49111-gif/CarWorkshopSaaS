from fastapi import FastAPI, APIRouter, HTTPException, Request, Query
from fastapi.responses import Response
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import io
import csv
import re
import logging
import uuid
import httpx
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Literal
from datetime import datetime, timezone, timedelta

from services.whatsapp import (
    notify_status_change as wa_notify_status,
    notify_invoice as wa_notify_invoice,
    notify_payment as wa_notify_payment,
    send_notification as wa_send,
    EVENTS as WA_EVENTS,
)


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


# ----------------- Models -----------------
class SessionRequest(BaseModel):
    session_id: str


class UserOut(BaseModel):
    user_id: str
    email: str
    name: str
    picture: Optional[str] = None
    role: str = "owner"


class JobPhotos(BaseModel):
    front: Optional[str] = None  # base64 data URI
    back: Optional[str] = None
    left: Optional[str] = None
    right: Optional[str] = None


class SparePart(BaseModel):
    name: str
    quantity: int = 1
    price: Optional[float] = None
    status: Literal["pending", "ordered", "installed"] = "pending"


JOB_STATUSES = [
    "vehicle_received",
    "inspection",
    "approval_pending",
    "repair_started",
    "quality_check",
    "ready_for_delivery",
    "delivered",
]
LEGACY_STATUS_MAP = {
    "pending": "vehicle_received",
    "in_progress": "repair_started",
    "completed": "delivered",
}


def normalize_status(s: Optional[str]) -> str:
    if not s:
        return "vehicle_received"
    if s in JOB_STATUSES:
        return s
    return LEGACY_STATUS_MAP.get(s, "vehicle_received")


class StatusHistoryEntry(BaseModel):
    status: str
    changed_at: str
    changed_by: str
    changed_by_name: Optional[str] = None
    note: Optional[str] = None


class JobCreate(BaseModel):
    customer_name: str
    customer_phone: Optional[str] = None
    car_name: str
    car_number: str
    model_year: Optional[str] = None
    reference: Optional[str] = None
    customer_problems: str
    photos: Optional[JobPhotos] = None
    odometer_km: Optional[int] = None
    assigned_service_advisor: Optional[str] = None


class JobUpdate(BaseModel):
    status: Optional[str] = None
    status_note: Optional[str] = None
    mechanic_findings: Optional[str] = None
    spare_parts: Optional[List[SparePart]] = None
    assigned_mechanic: Optional[str] = None
    assigned_service_advisor: Optional[str] = None
    odometer_km: Optional[int] = None
    labour_charges: Optional[float] = None
    discount: Optional[float] = None
    gst_rate: Optional[float] = None
    payment_status: Optional[Literal["unpaid", "partial", "paid"]] = None
    payment_note: Optional[str] = None


class Job(BaseModel):
    job_id: str
    job_card_no: Optional[str] = None
    customer_name: str
    customer_phone: Optional[str] = None
    car_name: str
    car_number: str
    model_year: Optional[str] = None
    odometer_km: Optional[int] = None
    reference: Optional[str] = None
    customer_problems: str
    mechanic_findings: Optional[str] = None
    spare_parts: List[SparePart] = []
    photos: JobPhotos = JobPhotos()
    status: str = "vehicle_received"
    status_history: List[StatusHistoryEntry] = []
    assigned_mechanic: Optional[str] = None
    assigned_service_advisor: Optional[str] = None
    labour_charges: float = 0
    discount: float = 0
    gst_rate: float = 18
    gst_amount: float = 0
    parts_total: float = 0
    total_amount: float = 0
    payment_status: str = "unpaid"
    estimated_cost: Optional[float] = None
    created_by: str
    created_at: str
    updated_at: str


# ----------------- Inventory & Billing Models -----------------
class PartBase(BaseModel):
    part_number: str
    name: str
    category: Optional[str] = None
    brand: Optional[str] = None
    supplier: Optional[str] = None
    purchase_price: Optional[float] = None
    mrp: Optional[float] = None
    gst: Optional[float] = None
    quantity: int = 0
    reserved_quantity: int = 0
    minimum_stock: Optional[int] = None
    warehouse_location: Optional[str] = None


class PartCreate(PartBase):
    pass


class PartOut(PartBase):
    part_id: str


class StockMove(BaseModel):
    quantity: int
    reason: Optional[str] = None


class InventoryTransaction(BaseModel):
    part_id: str
    change: int
    type: Literal["purchase", "usage", "return", "adjustment", "reservation"]
    reference: Optional[str] = None
    created_by: Optional[str] = None
    created_at: str = Field(default_factory=_now_iso)


class InvoiceLine(BaseModel):
    description: str
    quantity: int = 1
    unit_price: float
    gst_percent: Optional[float] = 0


class InvoiceCreate(BaseModel):
    job_id: str
    customer_id: Optional[str] = None
    lines: List[InvoiceLine]
    labour_amount: Optional[float] = 0
    discount: Optional[float] = 0


class InvoiceOut(BaseModel):
    invoice_id: str
    invoice_number: str
    job_id: str
    customer_id: Optional[str]
    total_amount: float
    status: str
    created_at: str


class AuthLogin(BaseModel):
    email: Optional[str] = None


class TokenRefresh(BaseModel):
    refreshToken: str


# ----------------- Auth helpers -----------------
async def get_current_user(request: Request) -> dict:
    auth_header = request.headers.get("Authorization", "")
    if not auth_header.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing bearer token")
    token = auth_header.split(" ", 1)[1]
    session = await db.user_sessions.find_one({"session_token": token}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=401, detail="Invalid session")
    exp = session.get("expires_at")
    if isinstance(exp, datetime):
        if exp.tzinfo is None:
            exp = exp.replace(tzinfo=timezone.utc)
        if exp < datetime.now(timezone.utc):
            raise HTTPException(status_code=401, detail="Session expired")
    user = await db.users.find_one({"user_id": session["user_id"]}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    return user


def require_roles(user: dict, allowed: List[str]) -> None:
    if not user:
        raise HTTPException(status_code=401, detail="Unauthorized")
    if user.get("role") not in allowed:
        raise HTTPException(status_code=403, detail="Forbidden")


async def _find_part_by_name_or_number(name: Optional[str], part_number: Optional[str]):
    q = {}
    if part_number:
        q["part_number"] = part_number
    elif name:
        q["name"] = name
    else:
        return None
    return await db.parts.find_one(q, {"_id": 0})



# ----------------- Auth Routes -----------------
@api_router.post("/auth/session")
async def create_session(payload: SessionRequest):
    """Exchange Emergent session_id for an app session_token."""
    async with httpx.AsyncClient(timeout=20.0) as hc:
        resp = await hc.get(
            "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
            headers={"X-Session-ID": payload.session_id},
        )
    if resp.status_code != 200:
        raise HTTPException(status_code=401, detail="Invalid session_id")
    data = resp.json()
    email = data["email"]

    existing = await db.users.find_one({"email": email}, {"_id": 0})
    if existing:
        user_id = existing["user_id"]
        role = existing.get("role", "owner")
        await db.users.update_one(
            {"user_id": user_id},
            {"$set": {"name": data.get("name", existing.get("name")),
                      "picture": data.get("picture", existing.get("picture"))}},
        )
    else:
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        # first user becomes owner, rest mechanic
        count = await db.users.count_documents({})
        role = "owner" if count == 0 else "mechanic"
        await db.users.insert_one({
            "user_id": user_id,
            "email": email,
            "name": data.get("name", ""),
            "picture": data.get("picture", ""),
            "role": role,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })

    session_token = data["session_token"]
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    await db.user_sessions.update_one(
        {"session_token": session_token},
        {"$set": {
            "session_token": session_token,
            "user_id": user_id,
            "expires_at": expires_at,
            "created_at": datetime.now(timezone.utc),
        }},
        upsert=True,
    )
    return {
        "session_token": session_token,
        "user": {
            "user_id": user_id,
            "email": email,
            "name": data.get("name", ""),
            "picture": data.get("picture", ""),
            "role": role,
        },
    }


@api_router.get("/auth/me", response_model=UserOut)
async def auth_me(request: Request):
    user = await get_current_user(request)
    return UserOut(**user)


@api_router.post("/auth/logout")
async def auth_logout(request: Request):
    auth_header = request.headers.get("Authorization", "")
    if auth_header.startswith("Bearer "):
        token = auth_header.split(" ", 1)[1]
        await db.user_sessions.delete_one({"session_token": token})
    return {"ok": True}


async def _create_session(user_id: str) -> str:
    session_token = f"session_{uuid.uuid4().hex[:16]}"
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    await db.user_sessions.update_one(
        {"session_token": session_token},
        {"$set": {
            "session_token": session_token,
            "user_id": user_id,
            "expires_at": expires_at,
            "created_at": datetime.now(timezone.utc),
        }},
        upsert=True,
    )
    return session_token


async def _ensure_demo_owner() -> dict:
    owner = await db.users.find_one({"role": "owner"}, {"_id": 0})
    if owner:
        return owner
    owner = {
        "user_id": "user_demo_owner",
        "email": "owner@demo.local",
        "name": "Demo Owner",
        "picture": "",
        "role": "owner",
        "created_at": _now_iso(),
    }
    await db.users.insert_one(owner)
    return owner


@api_router.post("/auth/login")
async def auth_login(payload: AuthLogin):
    user = None
    if payload.email:
        user = await db.users.find_one({"email": payload.email}, {"_id": 0})
    if not user:
        user = await _ensure_demo_owner()
    session_token = await _create_session(user["user_id"])
    return {
        "accessToken": session_token,
        "refreshToken": session_token,
        "user": user,
    }


@api_router.post("/auth/refresh")
async def auth_refresh(payload: TokenRefresh):
    session = await db.user_sessions.find_one({"session_token": payload.refreshToken}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=401, detail="Invalid refresh token")
    exp = session.get("expires_at")
    if isinstance(exp, datetime):
        if exp.tzinfo is None:
            exp = exp.replace(tzinfo=timezone.utc)
        if exp < datetime.now(timezone.utc):
            raise HTTPException(status_code=401, detail="Refresh token expired")
    new_token = await _create_session(session["user_id"])
    user = await db.users.find_one({"user_id": session["user_id"]}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    await db.user_sessions.delete_one({"session_token": payload.refreshToken})
    return {
        "accessToken": new_token,
        "refreshToken": new_token,
        "user": user,
    }


# ----------------- Job Routes -----------------
def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


async def _next_job_card_number() -> str:
    now = datetime.now(timezone.utc)
    prefix = f"JC-{now.strftime('%y%m')}-"
    # Count jobs already using this prefix for a stable per-month counter
    existing = await db.jobs.count_documents({"job_card_no": {"$regex": f"^{prefix}"}})
    return f"{prefix}{existing + 1:04d}"


def _compute_totals(doc: dict) -> dict:
    parts_total = sum(
        (p.get("price") or 0) * (p.get("quantity") or 1)
        for p in (doc.get("spare_parts") or [])
    )
    labour = float(doc.get("labour_charges") or 0)
    discount = float(doc.get("discount") or 0)
    gst_rate = float(doc.get("gst_rate") or 0)
    sub = max(0.0, parts_total + labour - discount)
    gst_amount = round(sub * gst_rate / 100, 2)
    total = round(sub + gst_amount, 2)
    doc["parts_total"] = round(parts_total, 2)
    doc["gst_amount"] = gst_amount
    doc["total_amount"] = total
    return doc


@api_router.post("/jobs", response_model=Job)
async def create_job(payload: JobCreate, request: Request):
    user = await get_current_user(request)
    job_id = f"job_{uuid.uuid4().hex[:12]}"
    now = _now_iso()
    photos_dict = (payload.photos.model_dump() if payload.photos else JobPhotos().model_dump())

    photos_bytes = sum(len(v or "") for v in photos_dict.values())
    if photos_bytes > 12_000_000:
        raise HTTPException(
            status_code=413,
            detail=(
                f"Photos too large ({photos_bytes // 1024 // 1024} MB). "
                "Re-take photos using the app's camera which auto-compresses, "
                "or pick a smaller image from gallery."
            ),
        )

    settings = await db.settings.find_one({"_id": "workshop"}) or {}
    default_gst = float(settings.get("default_gst_rate", 18))

    initial_status = "vehicle_received"
    history = [{
        "status": initial_status,
        "changed_at": now,
        "changed_by": user["user_id"],
        "changed_by_name": user.get("name"),
        "note": "Vehicle received at workshop",
    }]

    doc = {
        "job_id": job_id,
        "job_card_no": await _next_job_card_number(),
        "customer_name": payload.customer_name.strip(),
        "customer_phone": (payload.customer_phone or "").strip() or None,
        "car_name": payload.car_name.strip(),
        "car_number": payload.car_number.upper().strip(),
        "model_year": payload.model_year,
        "odometer_km": payload.odometer_km,
        "reference": payload.reference,
        "customer_problems": payload.customer_problems,
        "mechanic_findings": None,
        "spare_parts": [],
        "photos": photos_dict,
        "status": initial_status,
        "status_history": history,
        "assigned_mechanic": None,
        "assigned_service_advisor": payload.assigned_service_advisor,
        "labour_charges": 0,
        "discount": 0,
        "gst_rate": default_gst,
        "gst_amount": 0,
        "parts_total": 0,
        "total_amount": 0,
        "payment_status": "unpaid",
        "estimated_cost": None,
        "created_by": user["user_id"],
        "created_at": now,
        "updated_at": now,
    }
    _compute_totals(doc)
    await db.jobs.insert_one(doc)
    doc.pop("_id", None)
    # Fire notification (mock provider by default; logs to backend log)
    try:
        wa_notify_status(doc, settings)
    except Exception as e:
        logger.warning("notify failed on create: %s", e)
    return Job(**doc)


@api_router.get("/jobs", response_model=List[Job])
async def list_jobs(
    request: Request,
    q: Optional[str] = Query(None, description="Search by car number or customer name"),
    status: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    limit: int = 200,
):
    await get_current_user(request)
    query: dict = {}
    if status and status != "all":
        query["status"] = status
    if q:
        query["$or"] = [
            {"car_number": {"$regex": q, "$options": "i"}},
            {"customer_name": {"$regex": q, "$options": "i"}},
            {"car_name": {"$regex": q, "$options": "i"}},
        ]
    if date_from or date_to:
        rng: dict = {}
        if date_from:
            rng["$gte"] = date_from
        if date_to:
            rng["$lte"] = date_to
        query["created_at"] = rng

    cursor = db.jobs.find(query, {"_id": 0}).sort("created_at", -1).limit(limit)
    items = await cursor.to_list(length=limit)
    return [Job(**i) for i in items]


@api_router.get("/jobs/export.csv")
async def export_jobs_csv(
    request: Request,
    status: Optional[str] = Query(None),
    q: Optional[str] = Query(None),
):
    """Export all jobs as CSV (photos excluded, only flagged)."""
    await get_current_user(request)
    query: dict = {}
    if status and status != "all":
        query["status"] = status
    if q:
        query["$or"] = [
            {"car_number": {"$regex": q, "$options": "i"}},
            {"customer_name": {"$regex": q, "$options": "i"}},
            {"car_name": {"$regex": q, "$options": "i"}},
        ]
    rows = await db.jobs.find(query, {"_id": 0}).sort("created_at", -1).to_list(length=10000)

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "job_id", "created_at", "updated_at", "status",
        "customer_name", "customer_phone", "reference",
        "car_name", "car_number", "model_year",
        "customer_problems", "mechanic_findings",
        "spare_parts_count", "spare_parts_total_price",
        "spare_parts_detail",
        "photos_count", "photo_front", "photo_back", "photo_left", "photo_right",
    ])
    for r in rows:
        parts = r.get("spare_parts") or []
        total_price = sum(
            (p.get("price") or 0) * (p.get("quantity") or 1) for p in parts
        )
        parts_detail = " | ".join(
            f"{p.get('name')} x{p.get('quantity',1)} ({p.get('status','pending')})"
            for p in parts
        )
        photos = r.get("photos") or {}
        photo_keys = ["front", "back", "left", "right"]
        photos_count = sum(1 for k in photo_keys if photos.get(k))
        writer.writerow([
            r.get("job_id", ""),
            r.get("created_at", ""),
            r.get("updated_at", ""),
            r.get("status", ""),
            r.get("customer_name", ""),
            r.get("customer_phone", "") or "",
            r.get("reference", "") or "",
            r.get("car_name", ""),
            r.get("car_number", ""),
            r.get("model_year", "") or "",
            r.get("customer_problems", "") or "",
            r.get("mechanic_findings", "") or "",
            len(parts),
            f"{total_price:.2f}" if total_price else "",
            parts_detail,
            photos_count,
            "yes" if photos.get("front") else "",
            "yes" if photos.get("back") else "",
            "yes" if photos.get("left") else "",
            "yes" if photos.get("right") else "",
        ])

    csv_bytes = buf.getvalue().encode("utf-8")
    filename = f"workshop_jobs_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.csv"
    return Response(
        content=csv_bytes,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api_router.get("/jobs/export.xlsx")
async def export_jobs_xlsx(
    request: Request,
    status: Optional[str] = Query(None),
    q: Optional[str] = Query(None),
):
    """Export all jobs as a real Excel (.xlsx) workbook."""
    await get_current_user(request)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    query: dict = {}
    if status and status != "all":
        query["status"] = status
    if q:
        query["$or"] = [
            {"car_number": {"$regex": q, "$options": "i"}},
            {"customer_name": {"$regex": q, "$options": "i"}},
            {"car_name": {"$regex": q, "$options": "i"}},
        ]
    rows = await db.jobs.find(query, {"_id": 0}).sort("created_at", -1).to_list(length=10000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Workshop Jobs"

    headers = [
        "Job ID", "Created At", "Updated At", "Status",
        "Customer Name", "Phone", "Reference",
        "Car Name", "Car Number", "Model Year",
        "Customer Problems", "Mechanic Findings",
        "Parts Count", "Parts Total ₹", "Parts Detail",
        "Photos Count", "Front", "Back", "Left", "Right",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="FFD600")
    header_font = Font(bold=True, color="000000")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for r in rows:
        parts = r.get("spare_parts") or []
        total_price = sum((p.get("price") or 0) * (p.get("quantity") or 1) for p in parts)
        parts_detail = " | ".join(
            f"{p.get('name')} x{p.get('quantity', 1)} ({p.get('status', 'pending')})" for p in parts
        )
        photos = r.get("photos") or {}
        photo_keys = ["front", "back", "left", "right"]
        photos_count = sum(1 for k in photo_keys if photos.get(k))
        ws.append([
            r.get("job_id", ""),
            r.get("created_at", ""),
            r.get("updated_at", ""),
            (r.get("status") or "").replace("_", " ").title(),
            r.get("customer_name", ""),
            r.get("customer_phone", "") or "",
            r.get("reference", "") or "",
            r.get("car_name", ""),
            r.get("car_number", ""),
            r.get("model_year", "") or "",
            r.get("customer_problems", "") or "",
            r.get("mechanic_findings", "") or "",
            len(parts),
            round(total_price, 2) if total_price else 0,
            parts_detail,
            photos_count,
            "Yes" if photos.get("front") else "",
            "Yes" if photos.get("back") else "",
            "Yes" if photos.get("left") else "",
            "Yes" if photos.get("right") else "",
        ])

    # Auto-size columns (cap at 50)
    widths = [12, 22, 22, 14, 22, 14, 16, 22, 14, 10, 40, 40, 10, 14, 50, 10, 8, 8, 8, 8]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx) if idx <= 26 else f"A{chr(64 + idx - 26)}"].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"workshop_jobs_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api_router.get("/jobs/{job_id}", response_model=Job)
async def get_job(job_id: str, request: Request):
    await get_current_user(request)
    doc = await db.jobs.find_one({"job_id": job_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Job not found")
    return Job(**doc)


@api_router.patch("/jobs/{job_id}", response_model=Job)
async def update_job(job_id: str, payload: JobUpdate, request: Request):
    user = await get_current_user(request)
    existing = await db.jobs.find_one({"job_id": job_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Job not found")

    update: dict = {"updated_at": _now_iso()}

    # Status change → log to history
    if payload.status is not None:
        raw = payload.status
        if raw not in JOB_STATUSES and raw not in LEGACY_STATUS_MAP:
            raise HTTPException(status_code=400, detail="Invalid status")
        new_status = normalize_status(raw)
        if new_status != normalize_status(existing.get("status")):
            hist_entry = {
                "status": new_status,
                "changed_at": _now_iso(),
                "changed_by": user["user_id"],
                "changed_by_name": user.get("name"),
                "note": payload.status_note or None,
            }
            await db.jobs.update_one({"job_id": job_id}, {"$push": {"status_history": hist_entry}})
        update["status"] = new_status

    if payload.mechanic_findings is not None:
        update["mechanic_findings"] = payload.mechanic_findings
    if payload.assigned_mechanic is not None:
        update["assigned_mechanic"] = payload.assigned_mechanic
    if payload.assigned_service_advisor is not None:
        update["assigned_service_advisor"] = payload.assigned_service_advisor
    if payload.odometer_km is not None:
        update["odometer_km"] = payload.odometer_km
    if payload.labour_charges is not None:
        update["labour_charges"] = float(payload.labour_charges)
    if payload.discount is not None:
        update["discount"] = float(payload.discount)
    if payload.gst_rate is not None:
        update["gst_rate"] = float(payload.gst_rate)
    if payload.payment_status is not None:
        update["payment_status"] = payload.payment_status

    # Spare parts handling: persist list and adjust inventory for installed parts
    if payload.spare_parts is not None:
        parts_list = [sp.model_dump() for sp in payload.spare_parts]
        update["spare_parts"] = parts_list
        # For each part marked installed, try to deduct stock (existing behaviour)
        for p in parts_list:
            try:
                if (p.get("status") or "") == "installed":
                    candidate = await _find_part_by_name_or_number(p.get("name"), p.get("part_number"))
                    if candidate:
                        current = int(candidate.get("quantity", 0))
                        req_qty = int(p.get("quantity", 1))
                        if current < req_qty:
                            raise HTTPException(status_code=400, detail=f"Insufficient stock for {p.get('name')}")
                        await db.parts.update_one({"part_id": candidate.get("part_id")}, {"$inc": {"quantity": -req_qty}})
                        tx = {
                            "tx_id": f"tx_{uuid.uuid4().hex[:12]}",
                            "part_id": candidate.get("part_id"),
                            "change": -req_qty,
                            "type": "usage",
                            "reference": job_id,
                            "created_by": user.get("user_id"),
                            "created_at": _now_iso(),
                        }
                        await db.inventory_transactions.insert_one(tx)
            except HTTPException:
                raise
            except Exception:
                logger.exception("inventory adjust failed for part %s", p.get("name"))

    await db.jobs.update_one({"job_id": job_id}, {"$set": update})
    # Recompute totals with the merged doc
    doc = await db.jobs.find_one({"job_id": job_id}, {"_id": 0})
    doc = _compute_totals(doc)
    await db.jobs.update_one(
        {"job_id": job_id},
        {"$set": {
            "parts_total": doc["parts_total"],
            "gst_amount": doc["gst_amount"],
            "total_amount": doc["total_amount"],
        }},
    )
    # Lifecycle notifications
    try:
        settings = await db.settings.find_one({"_id": "workshop"}) or {}
        if payload.status is not None:
            wa_notify_status(doc, settings)
        if payload.payment_status == "paid":
            wa_notify_payment(doc, settings)
    except Exception as e:
        logger.warning("notify failed on update: %s", e)
    return Job(**doc)


# ----------------- Parts / Inventory -----------------
@api_router.post("/parts", response_model=PartOut)
async def create_part(payload: PartCreate, request: Request):
    user = await get_current_user(request)
    require_roles(user, ["owner", "manager", "accountant"])  # case-insensitive roles in DB may differ
    part_id = f"part_{uuid.uuid4().hex[:12]}"
    doc = {"part_id": part_id, **payload.model_dump(), "created_at": _now_iso(), "updated_at": _now_iso()}
    await db.parts.insert_one(doc)
    return PartOut(**doc)


@api_router.get("/parts")
async def list_parts(request: Request, q: Optional[str] = Query(None)):
    await get_current_user(request)
    query = {}
    if q:
        query["$or"] = [{"name": {"$regex": q, "$options": "i"}}, {"part_number": {"$regex": q, "$options": "i"}}]
    rows = await db.parts.find(query, {"_id": 0}).sort("name", 1).to_list(length=1000)
    return rows


@api_router.get("/parts/low-stock")
async def low_stock(request: Request):
    await get_current_user(request)
    rows = await db.parts.find(
        {"$expr": {"$lte": ["$quantity", {"$ifNull": ["$minimum_stock", 0]}]},
         "minimum_stock": {"$gt": 0}},
        {"_id": 0},
    ).sort("quantity", 1).to_list(length=500)
    return rows


@api_router.get("/parts/summary")
async def parts_summary(request: Request):
    await get_current_user(request)
    total_parts = await db.parts.count_documents({})
    low_count = await db.parts.count_documents(
        {"$expr": {"$lte": ["$quantity", {"$ifNull": ["$minimum_stock", 0]}]},
         "minimum_stock": {"$gt": 0}}
    )
    all_parts = await db.parts.find({}, {"_id": 0}).to_list(length=10000)
    inventory_value = sum(
        (p.get("purchase_price") or 0) * (p.get("quantity") or 0) for p in all_parts
    )
    retail_value = sum(
        (p.get("mrp") or 0) * (p.get("quantity") or 0) for p in all_parts
    )
    return {
        "total_parts": total_parts,
        "low_stock_count": low_count,
        "inventory_value": round(inventory_value, 2),
        "retail_value": round(retail_value, 2),
        "potential_margin": round(retail_value - inventory_value, 2),
    }


@api_router.get("/parts/export.xlsx")
async def export_parts_xlsx(request: Request):
    await get_current_user(request)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    rows = await db.parts.find({}, {"_id": 0}).sort("name", 1).to_list(length=10000)
    wb = Workbook()
    ws = wb.active
    ws.title = "Parts Inventory"
    headers = [
        "Part Number", "Name", "Category", "Brand", "Supplier",
        "Purchase Price", "MRP", "GST %", "Quantity", "Minimum Stock",
        "Warehouse Location",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="FFD600")
    header_font = Font(bold=True, color="000000")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left")
    for r in rows:
        ws.append([
            r.get("part_number", ""), r.get("name", ""), r.get("category", "") or "",
            r.get("brand", "") or "", r.get("supplier", "") or "",
            r.get("purchase_price") or 0, r.get("mrp") or 0, r.get("gst") or 0,
            r.get("quantity") or 0, r.get("minimum_stock") or 0,
            r.get("warehouse_location", "") or "",
        ])
    ws.freeze_panes = "A2"
    for idx, w in enumerate([14, 30, 16, 14, 16, 12, 10, 8, 10, 12, 16], start=1):
        ws.column_dimensions[chr(64 + idx)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"parts_inventory_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )


class BulkPartsImport(BaseModel):
    parts: List[PartCreate]


@api_router.post("/parts/import")
async def import_parts(payload: BulkPartsImport, request: Request):
    user = await get_current_user(request)
    if user.get("role") not in ("owner", "manager"):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    inserted, updated = 0, 0
    for p in payload.parts:
        data = p.model_dump()
        existing = await db.parts.find_one({"part_number": data["part_number"]}, {"part_id": 1})
        if existing:
            await db.parts.update_one(
                {"part_id": existing["part_id"]},
                {"$set": {**data, "updated_at": _now_iso()}},
            )
            updated += 1
        else:
            data.update({
                "part_id": f"part_{uuid.uuid4().hex[:12]}",
                "created_at": _now_iso(),
                "updated_at": _now_iso(),
            })
            await db.parts.insert_one(data)
            inserted += 1
    return {"ok": True, "inserted": inserted, "updated": updated}


class PartUpdate(BaseModel):
    part_number: Optional[str] = None
    name: Optional[str] = None
    category: Optional[str] = None
    brand: Optional[str] = None
    supplier: Optional[str] = None
    purchase_price: Optional[float] = None
    mrp: Optional[float] = None
    gst: Optional[float] = None
    quantity: Optional[int] = None
    minimum_stock: Optional[int] = None
    warehouse_location: Optional[str] = None


@api_router.get("/parts/{part_id}")
async def get_part(part_id: str, request: Request):
    await get_current_user(request)
    doc = await db.parts.find_one({"part_id": part_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Part not found")
    return doc


@api_router.patch("/parts/{part_id}", response_model=PartOut)
async def update_part(part_id: str, payload: PartUpdate, request: Request):
    user = await get_current_user(request)
    if user.get("role") not in ("owner", "manager", "accountant"):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    update = {k: v for k, v in payload.model_dump().items() if v is not None}
    if not update:
        raise HTTPException(status_code=400, detail="No fields to update")
    update["updated_at"] = _now_iso()
    res = await db.parts.update_one({"part_id": part_id}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Part not found")
    doc = await db.parts.find_one({"part_id": part_id}, {"_id": 0})
    return PartOut(**doc)


@api_router.delete("/parts/{part_id}")
async def delete_part(part_id: str, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "owner":
        raise HTTPException(status_code=403, detail="Only owners can delete parts")
    res = await db.parts.delete_one({"part_id": part_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Part not found")
    return {"ok": True}


@api_router.post("/parts/{part_id}/stock-move")
async def parts_stock_move(part_id: str, payload: StockMove, request: Request):
    user = await get_current_user(request)
    require_roles(user, ["owner", "manager", "accountant"])
    part = await db.parts.find_one({"part_id": part_id}, {"_id": 0})
    if not part:
        raise HTTPException(status_code=404, detail="Part not found")
    new_qty = (part.get("quantity") or 0) + int(payload.quantity)
    if new_qty < 0:
        raise HTTPException(status_code=400, detail="Insufficient stock")
    await db.parts.update_one({"part_id": part_id}, {"$set": {"quantity": new_qty, "updated_at": _now_iso()}})
    tx = {
        "tx_id": f"tx_{uuid.uuid4().hex[:12]}",
        "part_id": part_id,
        "change": int(payload.quantity),
        "type": "adjustment",
        "reference": payload.reason,
        "created_by": user.get("user_id"),
        "created_at": _now_iso(),
    }
    await db.inventory_transactions.insert_one(tx)
    return {"ok": True, "quantity": new_qty}


# ----------------- Invoices -----------------
@api_router.post("/invoices", response_model=InvoiceOut)
async def create_invoice(payload: InvoiceCreate, request: Request):
    user = await get_current_user(request)
    require_roles(user, ["owner", "accountant"])
    # Basic aggregation
    subtotal = 0.0
    for l in payload.lines:
        subtotal += l.quantity * l.unit_price
    subtotal += payload.labour_amount or 0
    total = subtotal - (payload.discount or 0)
    invoice_id = f"inv_{uuid.uuid4().hex[:12]}"
    invoice_number = f"INV{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')[:14]}"
    doc = {
        "invoice_id": invoice_id,
        "invoice_number": invoice_number,
        "job_id": payload.job_id,
        "customer_id": payload.customer_id,
        "lines": [l.model_dump() for l in payload.lines],
        "labour_amount": payload.labour_amount or 0,
        "discount": payload.discount or 0,
        "total_amount": total,
        "status": "draft",
        "created_at": _now_iso(),
        "created_by": user.get("user_id"),
    }
    await db.invoices.insert_one(doc)
    return InvoiceOut(**{
        "invoice_id": invoice_id,
        "invoice_number": invoice_number,
        "job_id": payload.job_id,
        "customer_id": payload.customer_id,
        "total_amount": total,
        "status": "draft",
        "created_at": doc["created_at"],
    })


@api_router.get("/invoices/{invoice_id}")
async def get_invoice(invoice_id: str, request: Request):
    await get_current_user(request)
    doc = await db.invoices.find_one({"invoice_id": invoice_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return doc


@api_router.delete("/jobs/{job_id}")
async def delete_job(job_id: str, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "owner":
        raise HTTPException(status_code=403, detail="Only owners can delete jobs")
    res = await db.jobs.delete_one({"job_id": job_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Job not found")
    return {"ok": True}


@api_router.get("/stats")
async def get_stats(request: Request):
    await get_current_user(request)
    pipeline = [{"$group": {"_id": "$status", "count": {"$sum": 1}}}]
    rows = await db.jobs.aggregate(pipeline).to_list(length=20)
    counts = {r["_id"]: r["count"] for r in rows}
    total = sum(counts.values())

    def buckets(names):
        return sum(counts.get(n, 0) for n in names)

    open_count = buckets(["vehicle_received", "inspection", "approval_pending"])
    working = buckets(["repair_started", "quality_check"])
    ready = counts.get("ready_for_delivery", 0)
    delivered = counts.get("delivered", 0)

    recent_cursor = db.jobs.find({}, {"_id": 0}).sort("created_at", -1).limit(5)
    recent = await recent_cursor.to_list(length=5)
    return {
        "total": total,
        # New buckets aligned to 7-state flow
        "open": open_count,
        "working": working,
        "ready": ready,
        "delivered": delivered,
        # Legacy keys — kept for backward compat with prior clients
        "pending": open_count,
        "in_progress": working,
        "completed": delivered,
        # Full per-status breakdown for advanced dashboards
        "by_status": {s: counts.get(s, 0) for s in JOB_STATUSES},
        "recent": recent,
    }


@api_router.get("/analytics")
async def get_analytics(request: Request):
    """Founder-level KPIs: trends, brands, issues, references, revenue, turnaround."""
    await get_current_user(request)
    jobs = await db.jobs.find({}, {"_id": 0}).sort("created_at", -1).to_list(length=10000)
    now = datetime.now(timezone.utc)

    # ----- Status counts (7-state flow + legacy aliases for old dashboards)
    status_counts = {s: 0 for s in JOB_STATUSES}
    for j in jobs:
        s = normalize_status(j.get("status"))
        if s in status_counts:
            status_counts[s] += 1
    # Convenience aggregates for the frontend
    status_counts["_pending"] = (
        status_counts["vehicle_received"] + status_counts["inspection"] + status_counts["approval_pending"]
    )
    status_counts["_working"] = status_counts["repair_started"] + status_counts["quality_check"]
    status_counts["_delivered"] = status_counts["delivered"]
    # Legacy keys
    status_counts["pending"] = status_counts["_pending"]
    status_counts["in_progress"] = status_counts["_working"]
    status_counts["completed"] = status_counts["_delivered"]

    # ----- Time windows (7d / 30d)
    def parse_dt(s):
        try:
            d = datetime.fromisoformat(s.replace("Z", "+00:00"))
            return d if d.tzinfo else d.replace(tzinfo=timezone.utc)
        except Exception:
            return None

    in_7d = 0
    in_30d = 0
    daily_buckets: dict = {}  # last 14 days
    for j in jobs:
        d = parse_dt(j.get("created_at", ""))
        if not d:
            continue
        delta_days = (now - d).days
        if delta_days <= 7:
            in_7d += 1
        if delta_days <= 30:
            in_30d += 1
        if delta_days <= 13:
            key = d.strftime("%Y-%m-%d")
            daily_buckets[key] = daily_buckets.get(key, 0) + 1

    # Last 14 daily series ordered ascending
    daily_series = []
    for i in range(13, -1, -1):
        day = (now - timedelta(days=i)).strftime("%Y-%m-%d")
        daily_series.append({"date": day, "count": daily_buckets.get(day, 0)})

    # ----- Brand breakdown (first token of car_name)
    brand_counts: dict = {}
    for j in jobs:
        car_name = (j.get("car_name") or "").strip()
        if not car_name:
            continue
        brand = car_name.split()[0].title()
        brand_counts[brand] = brand_counts.get(brand, 0) + 1
    brands = sorted(
        [{"label": k, "count": v} for k, v in brand_counts.items()],
        key=lambda x: x["count"],
        reverse=True,
    )[:8]

    # ----- Reference (referral source) breakdown
    ref_counts: dict = {}
    for j in jobs:
        ref = (j.get("reference") or "").strip()
        if not ref:
            ref = "Walk-in"
        ref_counts[ref] = ref_counts.get(ref, 0) + 1
    references = sorted(
        [{"label": k, "count": v} for k, v in ref_counts.items()],
        key=lambda x: x["count"],
        reverse=True,
    )[:8]

    # ----- Issue keywords (top words from customer_problems)
    STOP = {
        "the","a","an","is","it","of","to","in","on","and","or","with","my","for",
        "i","me","not","but","at","by","be","this","that","has","have","had","car",
        "from","as","so","very","when","its","while","also","please","sir","need",
        "needs","getting","got","showing","problem","problems","issue","issues",
        "vehicle","there","they","their","your","you","we","us","am","are","was",
        "were","do","does","done","just","only","than","then","into","out","up",
        "down","some","any","all","more","most","less","still","again",
    }
    word_counts: dict = {}
    for j in jobs:
        text = (j.get("customer_problems") or "").lower()
        for w in re.split(r"[^a-z]+", text):
            if len(w) < 3 or w in STOP:
                continue
            word_counts[w] = word_counts.get(w, 0) + 1
    issues = sorted(
        [{"label": k.title(), "count": v} for k, v in word_counts.items()],
        key=lambda x: x["count"],
        reverse=True,
    )[:8]

    # ----- Revenue (sum of spare parts qty*price across all jobs)
    revenue_total = 0.0
    revenue_completed = 0.0
    parts_total = 0
    for j in jobs:
        for p in j.get("spare_parts") or []:
            price = p.get("price") or 0
            qty = p.get("quantity") or 1
            revenue_total += price * qty
            parts_total += qty
            if normalize_status(j.get("status")) == "delivered":
                revenue_completed += price * qty

    # ----- Avg turnaround for completed jobs (hours)
    durations = []
    for j in jobs:
        if normalize_status(j.get("status")) != "delivered":
            continue
        c = parse_dt(j.get("created_at", ""))
        u = parse_dt(j.get("updated_at", ""))
        if c and u and u > c:
            durations.append((u - c).total_seconds() / 3600.0)
    avg_turnaround_hours = round(sum(durations) / len(durations), 1) if durations else None

    # ----- Top returning customers (by customer_name)
    customer_counts: dict = {}
    for j in jobs:
        name = (j.get("customer_name") or "").strip()
        if not name:
            continue
        customer_counts[name] = customer_counts.get(name, 0) + 1
    top_customers = sorted(
        [{"label": k, "count": v} for k, v in customer_counts.items() if v > 1],
        key=lambda x: x["count"],
        reverse=True,
    )[:6]

    # ----- Mechanic workload (assigned_mechanic)
    mech_counts: dict = {}
    for j in jobs:
        m = (j.get("assigned_mechanic") or "").strip()
        if not m:
            continue
        mech_counts[m] = mech_counts.get(m, 0) + 1
    mechanics = sorted(
        [{"label": k, "count": v} for k, v in mech_counts.items()],
        key=lambda x: x["count"],
        reverse=True,
    )[:6]

    # ----- Employee performance (who logged what)
    user_ids = list({j.get("created_by") for j in jobs if j.get("created_by")})
    name_map: dict = {}
    if user_ids:
        async for u in db.users.find({"user_id": {"$in": user_ids}}, {"_id": 0, "user_id": 1, "name": 1, "email": 1, "role": 1}):
            name_map[u["user_id"]] = {
                "name": u.get("name") or u.get("email") or u["user_id"],
                "role": u.get("role", "mechanic"),
            }
    emp_counts: dict = {}
    for j in jobs:
        uid = j.get("created_by")
        if not uid:
            continue
        d = emp_counts.setdefault(uid, {"intake": 0, "vehicle_received": 0, "inspection": 0, "approval_pending": 0, "repair_started": 0, "quality_check": 0, "ready_for_delivery": 0, "delivered": 0, "pending": 0, "in_progress": 0, "completed": 0, "today": 0, "week": 0, "month": 0})
        d["intake"] += 1
        s = normalize_status(j.get("status"))
        if s in d:
            d[s] += 1
        # Legacy fields
        if s in ("vehicle_received", "inspection", "approval_pending"):
            d["pending"] += 1
        elif s in ("repair_started", "quality_check"):
            d["in_progress"] += 1
        elif s == "delivered":
            d["completed"] += 1
        cdt = parse_dt(j.get("created_at", ""))
        if cdt:
            delta = (now - cdt).days
            if delta < 1:
                d["today"] += 1
            if delta <= 7:
                d["week"] += 1
            if delta <= 30:
                d["month"] += 1
    employees = []
    for uid, d in emp_counts.items():
        info = name_map.get(uid, {"name": uid, "role": "mechanic"})
        employees.append({
            "user_id": uid,
            "name": info["name"],
            "role": info["role"],
            **d,
        })
    employees.sort(key=lambda x: x["intake"], reverse=True)

    return {
        "total_jobs": len(jobs),
        "status_counts": status_counts,
        "intake_7d": in_7d,
        "intake_30d": in_30d,
        "daily_series": daily_series,
        "brands": brands,
        "references": references,
        "issues": issues,
        "revenue_total": round(revenue_total, 2),
        "revenue_completed": round(revenue_completed, 2),
        "parts_total": parts_total,
        "avg_turnaround_hours": avg_turnaround_hours,
        "completed_count": status_counts.get("completed", 0),
        "top_customers": top_customers,
        "mechanics": mechanics,
        "employees": employees,
        "unique_customers": len({(j.get("customer_name") or "").strip() for j in jobs if j.get("customer_name")}),
    }


# ----------------- Phase 1: Settings, Customers, Staff -----------------
ROLE_HIERARCHY = ["owner", "manager", "service_advisor", "mechanic", "accountant"]


class WorkshopSettings(BaseModel):
    workshop_name: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    pincode: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    gstin: Optional[str] = None
    default_gst_rate: float = 18
    upi_id: Optional[str] = None
    invoice_prefix: str = "INV"
    logo_base64: Optional[str] = None
    footer_note: Optional[str] = None


@api_router.get("/settings")
async def get_settings(request: Request):
    await get_current_user(request)
    doc = await db.settings.find_one({"_id": "workshop"}) or {}
    doc.pop("_id", None)
    return WorkshopSettings(**doc).model_dump()


@api_router.put("/settings")
async def update_settings(payload: WorkshopSettings, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "owner":
        raise HTTPException(status_code=403, detail="Only owners can update settings")
    data = payload.model_dump()
    await db.settings.update_one(
        {"_id": "workshop"},
        {"$set": data, "$setOnInsert": {"created_at": _now_iso()}},
        upsert=True,
    )
    return {"ok": True, **data}


@api_router.get("/customers")
async def list_customers(request: Request, q: Optional[str] = Query(None)):
    """Aggregated customer view derived from jobs collection."""
    await get_current_user(request)
    match: dict = {}
    if q:
        match["$or"] = [
            {"customer_name": {"$regex": q, "$options": "i"}},
            {"customer_phone": {"$regex": q, "$options": "i"}},
            {"car_number": {"$regex": q, "$options": "i"}},
        ]
    pipeline = [
        {"$match": match} if match else {"$match": {}},
        {"$group": {
            "_id": {
                "name": "$customer_name",
                "phone": {"$ifNull": ["$customer_phone", ""]},
            },
            "visits": {"$sum": 1},
            "lifetime_value": {"$sum": {"$ifNull": ["$total_amount", 0]}},
            "last_visit": {"$max": "$created_at"},
            "first_visit": {"$min": "$created_at"},
            "vehicles": {"$addToSet": {
                "car_number": "$car_number",
                "car_name": "$car_name",
                "model_year": "$model_year",
            }},
            "outstanding": {
                "$sum": {
                    "$cond": [
                        {"$ne": ["$payment_status", "paid"]},
                        {"$ifNull": ["$total_amount", 0]},
                        0,
                    ]
                }
            },
        }},
        {"$sort": {"last_visit": -1}},
        {"$limit": 1000},
    ]
    rows = await db.jobs.aggregate(pipeline).to_list(length=1000)
    result = []
    for r in rows:
        key = r["_id"]
        if not key.get("name"):
            continue
        result.append({
            "customer_name": key.get("name"),
            "customer_phone": key.get("phone") or None,
            "visits": r["visits"],
            "lifetime_value": round(r.get("lifetime_value") or 0, 2),
            "outstanding": round(r.get("outstanding") or 0, 2),
            "last_visit": r.get("last_visit"),
            "first_visit": r.get("first_visit"),
            "vehicles": r.get("vehicles", []),
        })
    return result


@api_router.get("/customers/{key}")
async def get_customer(key: str, request: Request):
    """key is customer_phone if it starts with +/digits, else customer_name."""
    await get_current_user(request)
    q: dict = (
        {"customer_phone": key}
        if key and (key[0] == "+" or key[0].isdigit())
        else {"customer_name": key}
    )
    jobs = await db.jobs.find(q, {"_id": 0}).sort("created_at", -1).to_list(length=200)
    if not jobs:
        raise HTTPException(status_code=404, detail="Customer not found")
    total = sum((j.get("total_amount") or 0) for j in jobs)
    outstanding = sum((j.get("total_amount") or 0) for j in jobs if (j.get("payment_status") or "unpaid") != "paid")
    vehicles = {}
    for j in jobs:
        cn = j.get("car_number")
        if cn and cn not in vehicles:
            vehicles[cn] = {
                "car_number": cn,
                "car_name": j.get("car_name"),
                "model_year": j.get("model_year"),
                "last_service": j.get("created_at"),
                "odometer_km": j.get("odometer_km"),
            }
    return {
        "customer_name": jobs[0].get("customer_name"),
        "customer_phone": jobs[0].get("customer_phone"),
        "visits": len(jobs),
        "lifetime_value": round(total, 2),
        "outstanding": round(outstanding, 2),
        "vehicles": list(vehicles.values()),
        "jobs": jobs,
    }


@api_router.get("/staff")
async def list_staff(request: Request):
    user = await get_current_user(request)
    if user.get("role") not in ("owner", "manager"):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    rows = await db.users.find({}, {"_id": 0}).sort("created_at", 1).to_list(length=1000)
    # Enrich each user with job counts
    for u in rows:
        u["intake"] = await db.jobs.count_documents({"created_by": u["user_id"]})
        u["completed"] = await db.jobs.count_documents({"created_by": u["user_id"], "status": "delivered"})
    return rows


class StaffUpdate(BaseModel):
    role: str


@api_router.patch("/staff/{user_id}")
async def update_staff(user_id: str, payload: StaffUpdate, request: Request):
    actor = await get_current_user(request)
    if actor.get("role") != "owner":
        raise HTTPException(status_code=403, detail="Only owners can change roles")
    if payload.role not in ROLE_HIERARCHY:
        raise HTTPException(status_code=400, detail="Invalid role")
    if user_id == actor["user_id"] and payload.role != "owner":
        raise HTTPException(status_code=400, detail="Owners cannot demote themselves")
    res = await db.users.update_one({"user_id": user_id}, {"$set": {"role": payload.role}})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"ok": True, "user_id": user_id, "role": payload.role}


# ----------------- Phase 2: PDF Invoice generation -----------------
@api_router.get("/jobs/{job_id}/invoice.pdf")
async def job_invoice_pdf(job_id: str, request: Request):
    await get_current_user(request)
    job = await db.jobs.find_one({"job_id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    settings = await db.settings.find_one({"_id": "workshop"}) or {}

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors as rc
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=15 * mm, bottomMargin=15 * mm,
                            leftMargin=15 * mm, rightMargin=15 * mm)
    styles = getSampleStyleSheet()
    styles["Normal"].fontSize = 9
    styles["Title"].fontSize = 20
    styles["Title"].textColor = rc.HexColor("#111111")

    elements = []

    # Workshop header
    workshop = settings.get("workshop_name") or "Multi-brand Car Workshop"
    address = " · ".join(filter(None, [
        settings.get("address"), settings.get("city"), settings.get("state"), settings.get("pincode")
    ]))
    contact = " · ".join(filter(None, [
        settings.get("phone"), settings.get("email"), settings.get("gstin") and f"GSTIN: {settings.get('gstin')}"
    ]))
    elements.append(Paragraph(f"<b>{workshop}</b>", styles["Title"]))
    if address:
        elements.append(Paragraph(address, styles["Normal"]))
    if contact:
        elements.append(Paragraph(contact, styles["Normal"]))
    elements.append(Spacer(1, 8))

    # Invoice meta table
    invoice_no = f"{settings.get('invoice_prefix', 'INV')}-{job.get('job_card_no', job['job_id'][-8:])}"
    meta_data = [
        ["Invoice #", invoice_no, "Job Card", job.get("job_card_no") or "-"],
        ["Date", (job.get("updated_at") or job.get("created_at") or "")[:10],
         "Status", (job.get("status") or "").replace("_", " ").upper()],
        ["Customer", job.get("customer_name") or "", "Phone", job.get("customer_phone") or "-"],
        ["Vehicle", f"{job.get('car_name') or ''} · {job.get('car_number') or ''}",
         "Model Year", job.get("model_year") or "-"],
    ]
    meta = Table(meta_data, colWidths=[24 * mm, 60 * mm, 24 * mm, 60 * mm])
    meta.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOX", (0, 0), (-1, -1), 0.6, rc.grey),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, rc.HexColor("#DDDDDD")),
        ("BACKGROUND", (0, 0), (0, -1), rc.HexColor("#F5F5F5")),
        ("BACKGROUND", (2, 0), (2, -1), rc.HexColor("#F5F5F5")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(meta)
    elements.append(Spacer(1, 12))

    # Parts + labour table
    line_data = [["#", "Description", "Qty", "Unit ₹", "Amount ₹"]]
    for i, p in enumerate(job.get("spare_parts") or [], start=1):
        qty = p.get("quantity") or 1
        price = p.get("price") or 0
        line_data.append([str(i), p.get("name") or "", str(qty), f"{price:.2f}", f"{price * qty:.2f}"])
    if job.get("labour_charges"):
        line_data.append(["-", "Labour Charges", "1", f"{job['labour_charges']:.2f}", f"{job['labour_charges']:.2f}"])
    if len(line_data) == 1:
        line_data.append(["-", "(no charges yet)", "-", "-", "-"])

    lines_table = Table(line_data, colWidths=[12 * mm, 90 * mm, 16 * mm, 26 * mm, 26 * mm])
    lines_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, 0), rc.HexColor("#FFD600")),
        ("BOX", (0, 0), (-1, -1), 0.6, rc.grey),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, rc.HexColor("#DDDDDD")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("PADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(lines_table)
    elements.append(Spacer(1, 10))

    # Totals
    parts_total = job.get("parts_total") or 0
    labour = job.get("labour_charges") or 0
    discount = job.get("discount") or 0
    gst_amt = job.get("gst_amount") or 0
    grand = job.get("total_amount") or 0
    totals_data = [
        ["Parts Subtotal", f"₹ {parts_total:.2f}"],
        ["Labour", f"₹ {labour:.2f}"],
        ["Discount", f"- ₹ {discount:.2f}"],
        [f"GST @ {job.get('gst_rate') or 0}%", f"₹ {gst_amt:.2f}"],
        ["GRAND TOTAL", f"₹ {grand:.2f}"],
    ]
    totals = Table(totals_data, colWidths=[130 * mm, 40 * mm])
    totals.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("PADDING", (0, 0), (-1, -1), 6),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("LINEABOVE", (0, -1), (-1, -1), 1, rc.black),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, -1), (-1, -1), 13),
        ("BACKGROUND", (0, -1), (-1, -1), rc.HexColor("#FFF7CC")),
    ]))
    elements.append(totals)
    elements.append(Spacer(1, 12))

    payment_status = (job.get("payment_status") or "unpaid").upper()
    upi = settings.get("upi_id")
    pay_lines = [f"<b>Payment Status:</b> {payment_status}"]
    if upi and grand > 0:
        pay_lines.append(
            f"Pay via UPI: <b>{upi}</b> · "
            f"upi://pay?pa={upi}&pn={workshop}&am={grand:.2f}&cu=INR"
        )
    for line in pay_lines:
        elements.append(Paragraph(line, styles["Normal"]))

    footer = settings.get("footer_note") or "Thank you for choosing us. All parts carry standard manufacturer warranty."
    elements.append(Spacer(1, 14))
    elements.append(Paragraph(f"<i>{footer}</i>", styles["Normal"]))

    doc.build(elements)
    buf.seek(0)
    fn = f"invoice_{invoice_no}.pdf"
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )


# ----------------- Phase 3: Notifications & Reminders -----------------
class ManualNotifyBody(BaseModel):
    event: str
    to_phone: Optional[str] = None
    context: Optional[dict] = None


@api_router.post("/notifications/send")
async def notifications_send(payload: ManualNotifyBody, request: Request):
    """Manually trigger a WhatsApp message (uses configured provider — default: console)."""
    user = await get_current_user(request)
    if user.get("role") not in ("owner", "manager", "service_advisor"):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    if payload.event not in WA_EVENTS:
        raise HTTPException(status_code=400, detail=f"Unknown event. Supported: {list(WA_EVENTS.keys())}")
    ctx = payload.context or {}
    settings = await db.settings.find_one({"_id": "workshop"}) or {}
    ctx = {**settings, **ctx}
    res = wa_send(payload.event, payload.to_phone, ctx)
    return {"ok": bool(res and res.ok), "provider": res.provider if res else "none", "message_id": res.message_id if res else None}


@api_router.get("/reminders/due")
async def reminders_due(
    request: Request,
    days_since: int = Query(180, description="Vehicles not serviced in N+ days"),
    km_since: int = Query(10000, description="Vehicles crossed since last odo by N+ km"),
):
    """Compute vehicles that are due for service based on time or mileage."""
    await get_current_user(request)
    cutoff = (datetime.now(timezone.utc) - timedelta(days=days_since)).isoformat()
    # Group by car_number, get latest entry per vehicle
    pipeline = [
        {"$sort": {"created_at": -1}},
        {"$group": {
            "_id": "$car_number",
            "customer_name": {"$first": "$customer_name"},
            "customer_phone": {"$first": "$customer_phone"},
            "car_name": {"$first": "$car_name"},
            "last_service": {"$first": "$created_at"},
            "last_odometer": {"$first": "$odometer_km"},
        }},
        {"$match": {"last_service": {"$lt": cutoff}}},
        {"$sort": {"last_service": 1}},
        {"$limit": 200},
    ]
    rows = await db.jobs.aggregate(pipeline).to_list(length=200)
    result = []
    for r in rows:
        if not r["_id"]:
            continue
        days = 0
        try:
            last = datetime.fromisoformat((r["last_service"] or "").replace("Z", "+00:00"))
            if last.tzinfo is None:
                last = last.replace(tzinfo=timezone.utc)
            days = (datetime.now(timezone.utc) - last).days
        except Exception:
            pass
        result.append({
            "car_number": r["_id"],
            "customer_name": r.get("customer_name"),
            "customer_phone": r.get("customer_phone"),
            "car_name": r.get("car_name"),
            "last_service": r.get("last_service"),
            "last_odometer_km": r.get("last_odometer"),
            "days_since_service": days,
            "reason": "time",
        })
    return {"count": len(result), "items": result}


@api_router.post("/reminders/send-due")
async def reminders_send_due(
    request: Request,
    days_since: int = Query(180),
):
    """Send WhatsApp reminders to all vehicles due for service (via configured provider)."""
    user = await get_current_user(request)
    if user.get("role") not in ("owner", "manager"):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    due = await reminders_due(request, days_since=days_since, km_since=0)
    settings = await db.settings.find_one({"_id": "workshop"}) or {}
    sent, skipped = 0, 0
    for item in due.get("items", []):
        if not item.get("customer_phone"):
            skipped += 1
            continue
        ctx = {**settings, **item, "workshop": settings.get("workshop_name") or "our workshop", "phone": settings.get("phone") or ""}
        res = wa_send("service_reminder_time", item["customer_phone"], ctx)
        if res and res.ok:
            sent += 1
    return {"ok": True, "sent": sent, "skipped": skipped, "total_due": len(due.get("items", []))}


@api_router.get("/notifications/events")
async def list_notification_events(request: Request):
    await get_current_user(request)
    return {"events": list(WA_EVENTS.keys()), "templates": WA_EVENTS}


@api_router.get("/")
async def root():
    return {"service": "WorkshopOps API", "status": "ok"}


# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def setup_indexes():
    try:
        await db.users.create_index("email", unique=True)
        await db.users.create_index("user_id", unique=True)
        await db.user_sessions.create_index("session_token", unique=True)
        await db.user_sessions.create_index("expires_at", expireAfterSeconds=0)
        await db.jobs.create_index("job_id", unique=True)
        await db.jobs.create_index("car_number")
        await db.jobs.create_index("status")
        await db.jobs.create_index([("created_at", -1)])
    except Exception as e:
        logger.warning(f"Index setup warning: {e}")


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
